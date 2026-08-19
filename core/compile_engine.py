from __future__ import annotations

from collections import defaultdict
from datetime import date, datetime, timedelta
from io import BytesIO
import hashlib
import json
import re

from openpyxl import load_workbook
from sqlalchemy import select, func

from .db import session_scope
from .models import (
    FPTK, Candidate, Application, PipelineEvent, PositionMaster, MappingUser,
    UploadBatch, SourceOwnership, SourcingSourceMeta, CompiledBlacklist, AuditLog,
    UnlinkedSourcing,
)
from .logic import normalize_text, normalize_availability, excel_date, to_int, to_float, derive_recruitment_model
from .compile_rules import (
    ValidationResult, validate_recruiter_workbook, validate_sto_workbook,
    clean_text, key_text, canonical_from, BUSINESS_UNITS, DIRECTORATES,
    REQUEST_REASONS, CATEGORIES, STATUSES, REASON_TO_CATEGORY,
)

STAGES = [
    ("Sourcing FL", ["Sourcing Freelance"], ["Tanggal Sourcing Freelance"], []),
    ("Sourcing HR", ["Sourcing HR"], ["Tanggal Sourcing", "Tanggal Sourcing HR"], ["Detail Keterangan Sourcing HR"]),
    ("Shortlist CV", ["Shortlist CV"], ["Tanggal Shortlist CV"], ["Detail Keterangan Shortlist CV"]),
    ("Psikotes", ["Psikotes"], ["Tanggal Psikotes / Cek psikotes", "Tanggal Psikotes"], ["Detail Keterangan Psikotes"]),
    ("HR Interview", ["HR Interview"], ["Tanggal HR Interview"], ["Detail Keterangan HR Interview"]),
    ("Technical Test / Case Study", ["Technical Test/ Case Study", "Technical Test / Case Study"], ["Tanggal Technical Test/ Case Study", "Tanggal Technical Test / Case Study"], ["Detail Keterangan Technical Test/ Case Study", "Detail Keterangan Technical Test / Case Study"]),
    ("Market Visit", ["Market Visit"], ["Tanggal Market Visit"], ["Detail Market Visit", "Detail Keterangan Market Visit"]),
    ("Shortcall User Interview", ["Shortcall User Interview"], ["Tanggal Shortcall User Interview"], ["Detail Keterangan Shortcall User Interview"]),
    ("User Interview", ["User Interview"], ["Tanggal User Interview"], ["Detail Keterangan User Interview"]),
    ("Panel Interview", ["Panel Interview"], ["Tanggal Panel Interview"], ["Detail Keterangan Panel Interview"]),
    ("Reference Check", ["Reference Check"], ["Tanggal Reference Check"], ["Detail Keterangan Reference Check"]),
    ("MCU", ["MCU"], ["Tanggal MCU"], ["Detail Keterangan MCU"]),
    ("Offering", ["Offering"], ["Tanggal Offering"], ["Detail Keterangan Offering"]),
    ("Day 1", ["Day 1"], ["Tanggal Day 1"], ["Detail Keterangan Day 1"]),
]
STAGE_RANK = {x[0]: i for i, x in enumerate(STAGES)}


def detect_source_type(filename: str, explicit: str | None = None) -> str:
    if explicit:
        e = explicit.strip().upper()
        if e in {"STO", "RECRUITER"}:
            return e
    name = clean_text(filename).upper()
    # Dedicated STO files may be renamed by browsers/Windows (e.g. STO(1).xlsm).
    if re.match(r"^STO(?:$|[\s_().-])", name) or "TULANG PUNGGUNG" in name:
        return "STO"
    return "RECRUITER"


def load_uploaded_workbook(data: bytes):
    # data_only=True mirrors compile behavior: read the displayed/calculated values.
    return load_workbook(BytesIO(data), data_only=True, keep_vba=False, read_only=False)


def _sha256(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def _weeknum_sunday(d: date | None) -> int | None:
    if not d:
        return None
    # Excel WEEKNUM(date, 1): week starts Sunday, Jan 1 belongs to week 1.
    return int(d.strftime("%U")) + 1


def _month_name(d: date | None) -> str | None:
    return d.strftime("%B") if d else None


def _level_number(level) -> int | None:
    m = re.search(r"([1-5])", clean_text(level))
    return int(m.group(1)) if m else None


def _sla_days(level) -> int:
    n = _level_number(level)
    if n == 5:
        return 60
    if n == 4:
        return 45
    return 30


def _sla_result(status: str, deadline: date | None, offering: date | None, today: date | None = None) -> str:
    today = today or date.today()
    if status == "Cancel":
        return "Cancel FPTK"
    if status == "Closed":
        if not offering:
            return "Data Incomplete - Offering Date Missing"
        if not deadline:
            return "Data Incomplete - SLA Deadline Missing"
        return "Closed lulus SLA" if offering <= deadline else "Closed tidak lulus SLA"
    if not deadline:
        return "Data Incomplete - SLA Deadline Missing"
    return "OP belum lewat SLA" if today <= deadline else "OP tidak lulus SLA"


def _filter_category(position: str | None, level: str | None, explicit=None) -> str | None:
    if clean_text(explicit):
        return clean_text(explicit)
    p = key_text(position)
    if p.startswith("cimory") or p.startswith("fresh"):
        return "CLAP FGDP"
    n = _level_number(level)
    if n in {1, 2}:
        return "Level 1-2"
    if n == 3:
        return "Level 3"
    if n == 4:
        return "Level 4"
    return None


def _bu_code_from_pic(kode_pic: str | None) -> str | None:
    k = re.sub(r"[^A-Z0-9]", "", clean_text(kode_pic).upper())
    if k.startswith("CORP") or k.startswith("HO"):
        return "HO"
    if k.startswith("JESS"):
        return "JESS"
    if k.startswith("CMD"):
        return "CMD"
    if k.startswith("BHC") or k.startswith("BCH"):
        return "BHC"
    if k.startswith("ARC"):
        return "ARC"
    for prefix in ["MS", "MP", "MB"]:
        if k.startswith(prefix):
            return prefix
    return None


def _json_load(text) -> dict:
    if not text:
        return {}
    try:
        x = json.loads(text)
        return x if isinstance(x, dict) else {}
    except Exception:
        return {}


def _nonblank(value) -> bool:
    return value is not None and clean_text(value) != ""


def _set_nonblank(obj, attr: str, value, converter=None):
    if not _nonblank(value):
        return False
    v = converter(value) if converter else value
    if v is None:
        return False
    if getattr(obj, attr) != v:
        setattr(obj, attr, v)
        return True
    return False


def _ownership(s, entity_type: str, entity_id: int):
    return s.scalar(select(SourceOwnership).where(SourceOwnership.entity_type == entity_type, SourceOwnership.entity_id == entity_id))


def _set_ownership(s, entity_type: str, entity_id: int, user: dict, filename: str, sheet: str, row: int, batch_id: int, preserve_existing_owner=False):
    own = _ownership(s, entity_type, entity_id)
    if own and preserve_existing_owner:
        return own
    if own is None:
        own = SourceOwnership(entity_type=entity_type, entity_id=entity_id)
        s.add(own)
    own.owner_user_id = user.get("id")
    own.owner_email = user.get("email") or "unknown"
    own.owner_name = user.get("display_name")
    own.source_file = filename
    own.source_sheet = sheet
    own.source_row = row
    own.upload_batch_id = batch_id
    return own


def can_edit_entity(user: dict, entity_type: str, entity_id: int) -> bool:
    if user.get("role") == "admin":
        return True
    with session_scope() as s:
        own = _ownership(s, entity_type, entity_id)
        return bool(own and own.owner_user_id == user.get("id"))


def validate_upload(data: bytes, filename: str, user: dict, explicit_source_type: str | None = None) -> tuple[ValidationResult, object]:
    wb = load_uploaded_workbook(data)
    source_type = detect_source_type(filename, explicit_source_type)
    result = validate_sto_workbook(wb, user) if source_type == "STO" else validate_recruiter_workbook(wb, user)
    return result, wb


def _create_batch(user: dict, filename: str, source_type: str, data: bytes, validation: ValidationResult) -> int:
    with session_scope() as s:
        obj = UploadBatch(
            owner_user_id=user.get("id"), owner_email=user.get("email") or "unknown",
            owner_name=user.get("display_name"), source_file=filename, source_type=source_type,
            file_sha256=_sha256(data), status="VALIDATED" if validation.passed else "FAILED_VALIDATION",
            fptk_rows=validation.counts.get("fptk", 0), sourcing_rows=validation.counts.get("sourcing", 0),
            blacklist_rows=validation.counts.get("blacklist", 0), validation_issue_count=len(validation.issues),
            validation_json=json.dumps(validation.issue_rows(), ensure_ascii=False, default=str),
        )
        s.add(obj); s.flush()
        return obj.id


def _canonical_bool_availability(v):
    s = key_text(v)
    if s in {"y", "v", "ya", "yes", "true", "1"}:
        return True
    if s in {"n", "x", "tidak", "no", "false", "0"}:
        return False
    return None


def _apply_standard_fptk(obj: FPTK, rec: dict, filename: str) -> bool:
    changed = False
    # Core validated values overwrite because they are mandatory/canonical.
    direct = {
        "kode_pic": rec.get("Kode PIC"), "fptk_date": rec.get("FPTK Date (Real)"),
        "fptk_date_code": rec.get("FPTK Date (Kode)"), "position": rec.get("Posisi"),
        "business_unit": rec.get("Business Unit"), "directorate": rec.get("Direktorat"),
        "division": rec.get("Divisi"), "department": rec.get("Department"),
        "level_fptk": rec.get("Level FPTK"), "level_number": rec.get("Level Number"),
        "request_reason": rec.get("Alasan Permintaan FPTK"), "category_fptk": rec.get("Category FPTK"),
        "pic_recruiter": rec.get("PIC Recruiter"), "vacancy": rec.get("Vacancy"),
        "status": rec.get("Status"), "cancel_date": rec.get("FPTK Cancel Date"),
        "offering_date": rec.get("Offering Date"),
    }
    for attr, val in direct.items():
        if getattr(obj, attr) != val:
            setattr(obj, attr, val); changed = True

    # Optional fields: blank NEVER overwrites an existing central value.
    optional = {
        "kode_angka": rec.get("Kode Angka"),
        "candidate_name": rec.get("Nama Kandidat"),
        "estimated_join": excel_date(rec.get("Estimasi Join")),
        "onboarding_location": rec.get("Lokasi Onboarding"),
        "website_upload_date": excel_date(rec.get("Tanggal Upload ke Website")),
        "manager_user": rec.get("User (Manager)"), "indirect_user": rec.get("Indirect User"),
        "work_location": rec.get("Lokasi Kerja"), "hr_location": rec.get("Lokasi HR"),
        "employee_status": rec.get("Status Karyawan"), "region": rec.get("Region"),
        "new_replacement": rec.get("NEW/REPLACEMENT"), "detail_category": rec.get("Detail Kategori"),
        "remarks": rec.get("Remarks"),
    }
    for attr, val in optional.items():
        changed |= _set_nonblank(obj, attr, val, normalize_text if isinstance(val, str) else None)

    laptop = rec.get("Kebutuhan Laptop")
    if _nonblank(laptop):
        b = _canonical_bool_availability(laptop)
        if b is not None and obj.laptop_needed != b:
            obj.laptop_needed = b; changed = True

    av = rec.get("FPTK Availability")
    if _nonblank(av):
        b = _canonical_bool_availability(av)
        if b is not None and obj.fptk_availability != b:
            obj.fptk_availability = b; changed = True

    filter_cat = _filter_category(obj.position, obj.level_fptk, rec.get("Filter Kategorisasi FPTK"))
    if filter_cat and obj.filter_category != filter_cat:
        obj.filter_category = filter_cat; changed = True

    bu_code = _bu_code_from_pic(obj.kode_pic)
    if bu_code and obj.bu_code != bu_code:
        obj.bu_code = bu_code; changed = True

    sla_days = _sla_days(obj.level_fptk)
    deadline = obj.fptk_date + timedelta(days=sla_days) if obj.fptk_date else None
    sla_result = _sla_result(obj.status, deadline, obj.offering_date)
    for attr, val in [("sla_days", sla_days), ("deadline_sla", deadline), ("sla_result", sla_result)]:
        if getattr(obj, attr) != val:
            setattr(obj, attr, val); changed = True

    extra = _json_load(obj.extra_json)
    extra.update({
        "week_fptk_date": _weeknum_sunday(obj.fptk_date_code or obj.fptk_date),
        "month_fptk_date": _month_name(obj.fptk_date_code or obj.fptk_date),
        "week_cancel_date": _weeknum_sunday(obj.cancel_date),
        "month_cancel_date": _month_name(obj.cancel_date),
        "week_offering_date": _weeknum_sunday(obj.offering_date),
        "month_offering": _month_name(obj.offering_date),
        "source_row": rec.get("__row__"),
        "source_type": "RECRUITER",
    })
    new_extra = json.dumps(extra, ensure_ascii=False, default=str)
    if obj.extra_json != new_extra:
        obj.extra_json = new_extra; changed = True
    obj.source_file = filename
    return changed


def _compile_fptk(s, validation: ValidationResult, user: dict, filename: str, batch_id: int, stats: dict):
    for rec in validation.normalized_fptk:
        code = rec.get("Kode Unik")
        if not code:
            continue
        matches = s.scalars(select(FPTK).where(func.lower(FPTK.kode_unik) == code.lower())).all()
        if len(matches) > 1:
            raise RuntimeError(f"Central Master duplicate Kode Unik: {code}. Validation seharusnya sudah memblok record ini.")
        obj = matches[0] if matches else None
        created = obj is None
        if created:
            obj = FPTK(kode_unik=code, position=rec.get("Posisi") or "")
            s.add(obj); s.flush()
        else:
            own = _ownership(s, "FPTK", obj.id)
            if own and own.owner_user_id not in (None, user.get("id")):
                raise PermissionError(f"Kode Unik {code} dimiliki user lain.")
        changed = _apply_standard_fptk(obj, rec, filename)
        s.flush()
        _set_ownership(s, "FPTK", obj.id, user, filename, "FPTK", rec.get("__row__"), batch_id)
        stats["fptk_inserted" if created else ("fptk_updated" if changed else "fptk_unchanged")] += 1
    s.flush()


def _find_header(ws, names: list[str], header_row=1):
    norm = {clean_text(ws.cell(header_row, c).value).casefold(): c for c in range(1, ws.max_column + 1)}
    for n in names:
        if n.casefold() in norm:
            return norm[n.casefold()]
    return None


def _compile_position_master(s, wb, stats):
    if "DB Kode Posisi" not in wb.sheetnames:
        return
    ws = wb["DB Kode Posisi"]
    header_row = 1
    cols = {
        "code": _find_header(ws, ["KODE", "Kode"], header_row),
        "position": _find_header(ws, ["POSITION", "Position", "Posisi"], header_row),
        "location": _find_header(ws, ["LOCATION", "Location"], header_row),
        "business_unit": _find_header(ws, ["BUSINESS UNIT", "Business Unit"], header_row),
        "division": _find_header(ws, ["DIVISION CHRIS", "Division", "Divisi"], header_row),
        "department": _find_header(ws, ["DEPARTMENT CHRIS", "Department", "Departemen"], header_row),
        "manager_user": _find_header(ws, ["USER (MANAGER)", "User (Manager)"], header_row),
        "indirect_user": _find_header(ws, ["INDIRECT USER", "Indirect User"], header_row),
        "directorate": _find_header(ws, ["DIRECTORATE", "Direktorat"], header_row),
        "year": _find_header(ws, ["YEAR", "Year"], header_row),
    }
    if not cols["position"]:
        return
    for r in range(2, ws.max_row + 1):
        pos = clean_text(ws.cell(r, cols["position"]).value)
        if not pos:
            continue
        def val(k):
            c = cols.get(k); return ws.cell(r, c).value if c else None
        loc, bu = clean_text(val("location")), clean_text(val("business_unit"))
        candidates = s.scalars(select(PositionMaster).where(
            func.lower(PositionMaster.position) == pos.lower(),
            func.lower(func.coalesce(PositionMaster.location, "")) == loc.lower(),
            func.lower(func.coalesce(PositionMaster.business_unit, "")) == bu.lower(),
        )).all()
        obj = candidates[-1] if candidates else None
        created = obj is None
        if created:
            obj = PositionMaster(position=pos, location=loc or None, business_unit=bu or None)
            s.add(obj)
        for attr in ["code", "division", "department", "manager_user", "indirect_user", "directorate"]:
            _set_nonblank(obj, attr, val(attr), normalize_text)
        yr = to_int(val("year"))
        if yr is not None:
            obj.year = yr
        stats["position_master_inserted" if created else "position_master_updated"] += 1


def _compile_mapping_user(s, wb, stats):
    if "Mapping User" not in wb.sheetnames:
        return
    ws = wb["Mapping User"]
    pos_col = _find_header(ws, ["POSITION", "Position"], 1)
    if not pos_col:
        return
    user_col = _find_header(ws, ["USER (MANAGER)", "User (Manager)"], 1)
    indirect_col = _find_header(ws, ["INDIRECT USER", "Indirect User"], 1)
    dir_col = _find_header(ws, ["DIRECTORATE", "Direktorat"], 1)
    for r in range(2, ws.max_row + 1):
        pos = clean_text(ws.cell(r, pos_col).value)
        if not pos:
            continue
        directorate = clean_text(ws.cell(r, dir_col).value) if dir_col else ""
        obj = s.scalar(select(MappingUser).where(
            func.lower(MappingUser.position) == pos.lower(),
            func.lower(func.coalesce(MappingUser.directorate, "")) == directorate.lower(),
        ))
        created = obj is None
        if created:
            obj = MappingUser(position=pos, directorate=directorate or None); s.add(obj)
        if user_col: _set_nonblank(obj, "manager_user", ws.cell(r, user_col).value, normalize_text)
        if indirect_col: _set_nonblank(obj, "indirect_user", ws.cell(r, indirect_col).value, normalize_text)
        stats["mapping_user_inserted" if created else "mapping_user_updated"] += 1


def _rec_get(rec: dict, *names):
    lookup = {clean_text(k).casefold(): v for k, v in rec.items() if not str(k).startswith("__")}
    for n in names:
        if n.casefold() in lookup:
            return lookup[n.casefold()]
    return None


def _phone_key(phone) -> str:
    return re.sub(r"\D", "", clean_text(phone))


def _candidate_identity(rec: dict, sourcing_date: date) -> str:
    email = clean_text(_rec_get(rec, "Email"))
    phone = _phone_key(_rec_get(rec, "Nomor HP", "No HP", "Phone"))
    name = clean_text(_rec_get(rec, "Nama", "Nama Kandidat"))
    if email:
        return "EMAIL|" + email.casefold()
    if phone:
        return "PHONE|" + phone
    if sourcing_date:
        return "DATE|" + sourcing_date.isoformat()
    return "NAME|" + name.casefold()


def _find_candidate(s, rec: dict, fptk_id: int, sourcing_date: date, owner_user_id: int | None):
    name = clean_text(_rec_get(rec, "Nama", "Nama Kandidat"))
    email = clean_text(_rec_get(rec, "Email"))
    phone = _phone_key(_rec_get(rec, "Nomor HP", "No HP", "Phone"))
    if email:
        c = s.scalar(select(Candidate).where(func.lower(Candidate.email) == email.lower()))
        if c:
            return c
    if phone:
        c = s.scalar(select(Candidate).where(func.lower(Candidate.name) == name.lower(), Candidate.phone == phone))
        if c:
            return c
    # Stable fallback for candidates without email/phone: same FPTK + owner + sourcing date + name.
    q = select(Candidate).join(Application, Application.candidate_id == Candidate.id).join(SourcingSourceMeta, SourcingSourceMeta.application_id == Application.id).where(
        Application.fptk_id == fptk_id,
        func.lower(Candidate.name) == name.lower(),
        SourcingSourceMeta.sourcing_date == sourcing_date,
        SourcingSourceMeta.owner_user_id == owner_user_id,
    )
    return s.scalar(q)


def _update_candidate_nonblank(c: Candidate, rec: dict):
    mappings = {
        "email": ["Email"], "phone": ["Nomor HP", "No HP", "Phone"], "domicile": ["Domisili"],
        "education_level": ["Jenjang Pendidikan"], "university_top10": ["Nama Universitas/Sekolah (TOP 10)"],
        "university_other": ["Nama Universitas/Sekolah Lainnya"], "major": ["Jurusan"],
        "last_position": ["Last Position"], "last_tenure": ["Last Tenure"], "last_company": ["Last Company"],
        "total_tenure": ["Total Tenure"], "university_tier": ["TIER Universitas"], "gpa_tier": ["TIER IPK"],
    }
    for attr, names in mappings.items():
        raw = _rec_get(rec, *names)
        if _nonblank(raw):
            val = _phone_key(raw) if attr == "phone" else normalize_text(raw)
            if val and getattr(c, attr) != val:
                setattr(c, attr, val)
    for attr, names, conv in [
        ("graduation_year", ["Tahun Lulus"], to_int), ("gpa", ["IPK"], to_float),
        ("english_score", ["Skor Bahasa Inggris"], to_float),
    ]:
        raw = _rec_get(rec, *names)
        if _nonblank(raw):
            val = conv(raw)
            if val is not None:
                setattr(c, attr, val)
    raw = _rec_get(rec, "Pernah di FMCG?")
    if _nonblank(raw):
        c.fmcg_experience = normalize_availability(raw, False)


def _event_value(rec: dict, aliases: list[str]):
    return _rec_get(rec, *aliases) if aliases else None


def _sourcing_identity_key(rec: dict, sourcing_date: date) -> str:
    code = clean_text(_rec_get(rec, "Kode Unik"))
    name = clean_text(_rec_get(rec, "Nama", "Nama Kandidat"))
    return f"{code}|{name.casefold()}|{_candidate_identity(rec, sourcing_date)}"


def _store_unlinked_sourcing(s, rec: dict, user: dict, filename: str, batch_id: int, stats: dict, reason: str = "FPTK_NOT_FOUND"):
    """Persist a DB Sourcing row without inventing a parent FPTK.

    Unlinked rows are intentionally kept outside Application/Pipeline, so the
    existing Funneling/Progress queries (which join Application -> FPTK) will not
    count them until a user fixes the Kode Unik and links the row to a real FPTK.
    """
    sourcing_date = rec.get("Sourcing Date")
    code = clean_text(_rec_get(rec, "Kode Unik"))
    name = clean_text(_rec_get(rec, "Nama", "Nama Kandidat"))
    if not code or not name or not sourcing_date:
        stats["sourcing_skipped_missing_key"] += 1
        return None

    identity_key = _sourcing_identity_key(rec, sourcing_date)
    obj = s.scalar(select(UnlinkedSourcing).where(
        UnlinkedSourcing.owner_user_id == user.get("id"),
        UnlinkedSourcing.source_identity_key == identity_key,
    ))
    created = obj is None
    if obj is None:
        obj = UnlinkedSourcing(
            kode_unik=code,
            candidate_name=name,
            sourcing_date=sourcing_date,
            source_identity_key=identity_key,
            owner_user_id=user.get("id"),
            owner_email=user.get("email") or "unknown",
        )
        s.add(obj)

    obj.kode_unik = code
    obj.candidate_name = name
    obj.sourcing_date = sourcing_date
    obj.position = normalize_text(_rec_get(rec, "Posisi"))
    obj.recruiter = normalize_text(_rec_get(rec, "Rekruter", "PIC Recruiter"))
    obj.owner_user_id = user.get("id")
    obj.owner_email = user.get("email") or "unknown"
    obj.owner_name = user.get("display_name")
    obj.source_file = filename
    obj.source_row = rec.get("__row__")
    obj.upload_batch_id = batch_id
    obj.reason = reason
    obj.status = "UNLINKED"
    obj.resolved_fptk_id = None
    obj.resolved_application_id = None
    obj.resolved_at = None
    obj.raw_json = json.dumps({k: v for k, v in rec.items() if not str(k).startswith("__")}, ensure_ascii=False, default=str)
    stats["sourcing_unlinked_inserted" if created else "sourcing_unlinked_updated"] += 1
    return obj


def _compile_sourcing_record(s, rec: dict, f: FPTK, user: dict, filename: str, batch_id: int, stats: dict):
    """Compile one already-linked DB Sourcing record into Candidate/Application/Pipeline."""
    sourcing_date = rec.get("Sourcing Date")
    code = clean_text(_rec_get(rec, "Kode Unik"))
    name = clean_text(_rec_get(rec, "Nama", "Nama Kandidat"))
    c = _find_candidate(s, rec, f.id, sourcing_date, user.get("id"))
    created_candidate = c is None
    if c is None:
        c = Candidate(name=name)
        s.add(c); s.flush()
    _update_candidate_nonblank(c, rec)
    stats["candidate_inserted" if created_candidate else "candidate_updated"] += 1

    app = s.scalar(select(Application).where(Application.fptk_id == f.id, Application.candidate_id == c.id))
    created_app = app is None
    if app is None:
        app = Application(fptk_id=f.id, candidate_id=c.id)
        s.add(app); s.flush()
    own = _ownership(s, "Application", app.id)
    if own and own.owner_user_id not in (None, user.get("id")):
        raise PermissionError(f"Application {name} / {code} dimiliki source user lain.")
    recruiter = clean_text(_rec_get(rec, "Rekruter", "PIC Recruiter")) or f.pic_recruiter
    model = clean_text(_rec_get(rec, "Model Rekrutmen")) or derive_recruitment_model(f.position, f.level_fptk, f.filter_category)
    if recruiter: app.recruiter = recruiter
    if model: app.model = model
    app.status = "ACTIVE"

    identity = _candidate_identity(rec, sourcing_date)
    meta = s.scalar(select(SourcingSourceMeta).where(SourcingSourceMeta.application_id == app.id))
    if meta is None:
        meta = SourcingSourceMeta(application_id=app.id, sourcing_date=sourcing_date, owner_email=user.get("email") or "unknown")
        s.add(meta)
    meta.sourcing_date = sourcing_date
    meta.source_identity_key = f"{code}|{name.casefold()}|{identity}"
    meta.source_file = filename
    meta.source_row = rec.get("__row__")
    meta.owner_user_id = user.get("id")
    meta.owner_email = user.get("email") or "unknown"
    meta.raw_json = json.dumps({k: v for k, v in rec.items() if not str(k).startswith("__")}, ensure_ascii=False, default=str)

    latest_stage = app.current_stage
    for stage, status_alias, date_alias, detail_alias in STAGES:
        result = normalize_text(_event_value(rec, status_alias))
        dt = excel_date(_event_value(rec, date_alias))
        detail = normalize_text(_event_value(rec, detail_alias))
        if not (result or dt or detail):
            continue
        ev = s.scalar(select(PipelineEvent).where(PipelineEvent.application_id == app.id, PipelineEvent.stage == stage, PipelineEvent.event_date == dt))
        if ev is None:
            ev = PipelineEvent(application_id=app.id, stage=stage, event_date=dt, created_by=user.get("email"))
            s.add(ev); stats["pipeline_event_inserted"] += 1
        else:
            stats["pipeline_event_updated"] += 1
        if result: ev.result = result
        if detail: ev.detail = detail
        if latest_stage is None or STAGE_RANK.get(stage, -1) >= STAGE_RANK.get(latest_stage, -1):
            latest_stage = stage
    app.current_stage = latest_stage
    s.flush()
    _set_ownership(s, "Application", app.id, user, filename, "DB Sourcing", rec.get("__row__"), batch_id)
    stats["application_inserted" if created_app else "application_updated"] += 1
    return app


def _compile_sourcing(s, validation: ValidationResult, user: dict, filename: str, batch_id: int, stats: dict):
    # Latest nonblank source value wins because records are processed in source row order.
    for rec in validation.normalized_sourcing:
        sourcing_date = rec.get("Sourcing Date")
        code = clean_text(_rec_get(rec, "Kode Unik"))
        name = clean_text(_rec_get(rec, "Nama", "Nama Kandidat"))
        if not code or not name or not sourcing_date:
            stats["sourcing_skipped_missing_key"] += 1
            continue

        fptk_matches = s.scalars(select(FPTK).where(func.lower(FPTK.kode_unik) == code.lower())).all()
        if len(fptk_matches) == 0:
            _store_unlinked_sourcing(s, rec, user, filename, batch_id, stats, "FPTK_NOT_FOUND")
            continue
        if len(fptk_matches) > 1:
            _store_unlinked_sourcing(s, rec, user, filename, batch_id, stats, "MULTIPLE_FPTK_MATCH")
            continue

        f = fptk_matches[0]
        _compile_sourcing_record(s, rec, f, user, filename, batch_id, stats)

        # If this exact sourcing record used to be unresolved, close the warning
        # automatically after a later upload provides a valid parent FPTK.
        identity_key = _sourcing_identity_key(rec, sourcing_date)
        unresolved = s.scalar(select(UnlinkedSourcing).where(
            UnlinkedSourcing.owner_user_id == user.get("id"),
            UnlinkedSourcing.source_identity_key == identity_key,
            UnlinkedSourcing.status == "UNLINKED",
        ))
        if unresolved:
            unresolved.status = "RESOLVED"
            unresolved.resolved_fptk_id = f.id
            unresolved.resolved_at = datetime.utcnow()
            stats["sourcing_unlinked_auto_resolved"] += 1


def resolve_unlinked_sourcing(unlinked_id: int, new_kode_unik: str, user: dict) -> dict:
    """Fix an Unlinked DB Sourcing row from the UI and link it to an existing FPTK."""
    new_code = clean_text(new_kode_unik)
    if not new_code:
        raise ValueError("Kode Unik baru wajib diisi.")

    stats = defaultdict(int)
    with session_scope() as s:
        item = s.get(UnlinkedSourcing, int(unlinked_id))
        if item is None or item.status != "UNLINKED":
            raise ValueError("Data Unlinked tidak ditemukan atau sudah diselesaikan.")
        if user.get("role") != "admin" and item.owner_user_id != user.get("id"):
            raise PermissionError("Hanya source owner atau Admin yang boleh memperbaiki data ini.")

        matches = s.scalars(select(FPTK).where(func.lower(FPTK.kode_unik) == new_code.lower())).all()
        if len(matches) == 0:
            raise ValueError(f"Kode Unik '{new_code}' belum ada di Central FPTK.")
        if len(matches) > 1:
            raise ValueError(f"Kode Unik '{new_code}' masih duplicate di Central FPTK. Minta Admin membereskan duplicate dulu.")
        f = matches[0]

        try:
            raw = json.loads(item.raw_json or "{}")
        except Exception:
            raw = {}
        raw["Kode Unik"] = new_code
        raw["Nama"] = raw.get("Nama") or item.candidate_name
        raw["Sourcing Date"] = item.sourcing_date
        raw["__row__"] = item.source_row

        app = _compile_sourcing_record(
            s, raw, f, user,
            item.source_file or "Unlinked Sourcing Fix",
            item.upload_batch_id or 0,
            stats,
        )

        item.kode_unik = new_code
        item.status = "RESOLVED"
        item.resolved_fptk_id = f.id
        item.resolved_application_id = app.id
        item.resolved_at = datetime.utcnow()
        item.reason = "RESOLVED_BY_USER"
        s.add(AuditLog(
            user_email=user.get("email"),
            action="RESOLVE_UNLINKED_SOURCING",
            entity_type="UnlinkedSourcing",
            entity_id=str(item.id),
            detail=json.dumps({"new_kode_unik": new_code, "fptk_id": f.id, "application_id": app.id}, ensure_ascii=False),
        ))

    return dict(stats)

def _compile_blacklist(s, validation: ValidationResult, user: dict, filename: str, batch_id: int, stats: dict):
    for rec in validation.normalized_blacklist:
        no = clean_text(rec.get("No"))
        name = clean_text(rec.get("Nama Kandidat"))
        if not no or not name:
            continue
        obj = s.scalar(select(CompiledBlacklist).where(CompiledBlacklist.owner_user_id == user.get("id"), CompiledBlacklist.source_no == no))
        created = obj is None
        if obj is None:
            obj = CompiledBlacklist(owner_user_id=user.get("id"), owner_email=user.get("email") or "unknown", source_no=no, candidate_name=name)
            s.add(obj); s.flush()
        obj.owner_email = user.get("email") or "unknown"
        obj.candidate_name = name
        # Blank never overwrites existing blacklist fields either.
        _set_nonblank(obj, "business_unit", rec.get("Business Unit"), normalize_text)
        _set_nonblank(obj, "position", rec.get("Posisi"), normalize_text)
        _set_nonblank(obj, "location", rec.get("Lokasi"), normalize_text)
        _set_nonblank(obj, "category", rec.get("Kategori"), normalize_text)
        _set_nonblank(obj, "reason", rec.get("Alasan Tidak Proceed"), normalize_text)
        _set_nonblank(obj, "pic_recruiter", rec.get("PIC Rekruter"), normalize_text)
        lu = excel_date(rec.get("Last Update"))
        if lu: obj.last_update = lu
        obj.source_file = filename; obj.source_row = rec.get("__row__")
        stats["blacklist_inserted" if created else "blacklist_updated"] += 1
        _set_ownership(s, "CompiledBlacklist", obj.id, user, filename, "Blacklist Candidate", rec.get("__row__"), batch_id)


def _sto_synthetic_code(user: dict, filename: str, row: int, position: str) -> str:
    base = f"{user.get('id')}|{filename}|{row}|{position}".encode()
    return "STO@" + hashlib.sha1(base).hexdigest()[:18].upper()


def _sto_match(s, rec: dict, filename: str, user: dict):
    pos = clean_text(rec.get("Posisi"))
    cand = clean_text(rec.get("Nama Kandidat"))
    code = clean_text(rec.get("Kode Unik"))
    # VBA priority: Position + Candidate, then Kode Unik, then source-row marker for blank candidate/code.
    if cand:
        matches = s.scalars(select(FPTK).where(func.lower(FPTK.position) == pos.lower(), func.lower(func.coalesce(FPTK.candidate_name, "")) == cand.lower())).all()
        if matches:
            return matches[0]
    if code:
        matches = s.scalars(select(FPTK).where(func.lower(FPTK.kode_unik) == code.lower())).all()
        if matches:
            return matches[0]
    if not cand and not code:
        # Find previously inserted STO row by its stable synthetic code.
        synthetic = _sto_synthetic_code(user, filename, rec.get("__row__"), pos)
        return s.scalar(select(FPTK).where(FPTK.kode_unik == synthetic))
    return None


def _compile_sto(s, validation: ValidationResult, user: dict, filename: str, batch_id: int, stats: dict):
    for rec in validation.normalized_fptk:
        pos = clean_text(rec.get("Posisi"))
        av = clean_text(rec.get("FPTK Availability")).upper()
        if not pos or av not in {"Y", "N"}:
            continue
        obj = _sto_match(s, rec, filename, user)
        created = obj is None
        if obj is None:
            code = clean_text(rec.get("Kode Unik")) or _sto_synthetic_code(user, filename, rec.get("__row__"), pos)
            obj = FPTK(kode_unik=code, position=pos)
            s.add(obj); s.flush()
            # Best-effort copy because standard validation is intentionally bypassed for STO.
            obj.kode_pic = normalize_text(rec.get("Kode PIC"))
            obj.kode_angka = normalize_text(rec.get("Kode Angka"))
            obj.fptk_date = excel_date(rec.get("FPTK Date (Real)"))
            obj.fptk_date_code = excel_date(rec.get("FPTK Date (Kode)"))
            obj.business_unit = normalize_text(rec.get("Business Unit"))
            obj.directorate = normalize_text(rec.get("Direktorat"))
            obj.division = normalize_text(rec.get("Divisi"))
            obj.department = normalize_text(rec.get("Department"))
            obj.level_fptk = normalize_text(rec.get("Level FPTK"))
            obj.level_number = to_int(rec.get("Level Number"), _level_number(obj.level_fptk))
            obj.request_reason = normalize_text(rec.get("Alasan Permintaan FPTK"))
            obj.category_fptk = normalize_text(rec.get("Category FPTK"))
            obj.pic_recruiter = normalize_text(rec.get("PIC Recruiter"))
            obj.vacancy = to_int(rec.get("Vacancy"), 1) or 1
            raw_status = clean_text(rec.get("Status"))
            obj.status = canonical_from(raw_status, STATUSES) or "OP"
            obj.cancel_date = excel_date(rec.get("FPTK Cancel Date"))
            obj.offering_date = excel_date(rec.get("Offering Date"))
            obj.candidate_name = normalize_text(rec.get("Nama Kandidat"))
            obj.work_location = normalize_text(rec.get("Lokasi Kerja"))
            obj.region = normalize_text(rec.get("Region"))
            obj.source_file = filename
        obj.filter_category = "STO"
        obj.fptk_availability = (av == "Y")
        obj.source_file = filename
        # Derived SLA when enough data is available; missing offering on Closed is explicitly incomplete.
        obj.sla_days = _sla_days(obj.level_fptk)
        obj.deadline_sla = obj.fptk_date + timedelta(days=obj.sla_days) if obj.fptk_date else None
        obj.sla_result = _sla_result(obj.status, obj.deadline_sla, obj.offering_date)
        extra = _json_load(obj.extra_json)
        extra.update({
            "source_type": "STO", "sto_source_row": rec.get("__row__"), "sto_availability": av,
            "sto_original_kode_unik": clean_text(rec.get("Kode Unik")) or None,
            "week_fptk_date": _weeknum_sunday(obj.fptk_date_code or obj.fptk_date),
            "month_fptk_date": _month_name(obj.fptk_date_code or obj.fptk_date),
            "week_cancel_date": _weeknum_sunday(obj.cancel_date), "month_cancel_date": _month_name(obj.cancel_date),
            "week_offering_date": _weeknum_sunday(obj.offering_date), "month_offering": _month_name(obj.offering_date),
        })
        obj.extra_json = json.dumps(extra, ensure_ascii=False, default=str)
        if created:
            _set_ownership(s, "FPTK", obj.id, user, filename, "FPTK", rec.get("__row__"), batch_id)
        # Existing recruiter FPTK keeps its original ownership; STO is a central sync overlay.
        stats["sto_inserted" if created else "sto_updated"] += 1


def compile_validated_workbook(wb, validation: ValidationResult, user: dict, filename: str, batch_id: int, progress_callback=None) -> dict:
    if not validation.passed:
        raise ValueError("Compile dibatalkan karena validation masih memiliki blocking error.")
    stats = defaultdict(int)
    source_type = validation.source_type
    def progress(stage, pct, message):
        if progress_callback:
            progress_callback(stage, pct, message)
    try:
        with session_scope() as s:
            if source_type == "STO":
                progress("STO", 20, "Sync Y/N availability dan STO categorization")
                _compile_sto(s, validation, user, filename, batch_id, stats)
            else:
                progress("FPTK", 15, "Compile FPTK")
                _compile_fptk(s, validation, user, filename, batch_id, stats)
                progress("Master Data", 35, "Reconcile DB Kode Posisi & Mapping User")
                _compile_position_master(s, wb, stats)
                _compile_mapping_user(s, wb, stats)
                progress("DB Sourcing", 55, "Compile candidate/application/pipeline")
                _compile_sourcing(s, validation, user, filename, batch_id, stats)
                progress("Blacklist", 80, "Compile Blacklist Candidate")
                _compile_blacklist(s, validation, user, filename, batch_id, stats)
            progress("Finalize", 95, "Commit ownership & audit")
            s.add(AuditLog(user_email=user.get("email"), action="COMPILE_UPLOAD", entity_type="UploadBatch", entity_id=str(batch_id), detail=json.dumps(dict(stats), ensure_ascii=False)))
        with session_scope() as s:
            batch = s.get(UploadBatch, batch_id)
            batch.status = "COMPILED"
            batch.compiled_at = datetime.utcnow()
            batch.compile_json = json.dumps(dict(stats), ensure_ascii=False)
        progress("Complete", 100, "Compile selesai")
        return dict(stats)
    except Exception as exc:
        with session_scope() as s:
            batch = s.get(UploadBatch, batch_id)
            if batch:
                batch.status = "FAILED_COMPILE"
                batch.compile_json = json.dumps({"error": str(exc)}, ensure_ascii=False)
        raise


def process_upload(data: bytes, filename: str, user: dict, explicit_source_type: str | None = None, progress_callback=None):
    validation, wb = validate_upload(data, filename, user, explicit_source_type)
    batch_id = _create_batch(user, filename, validation.source_type, data, validation)
    if not validation.passed:
        return {"batch_id": batch_id, "validation": validation, "compiled": False, "stats": {}}
    stats = compile_validated_workbook(wb, validation, user, filename, batch_id, progress_callback)
    return {"batch_id": batch_id, "validation": validation, "compiled": True, "stats": stats}
