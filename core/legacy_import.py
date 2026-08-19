from __future__ import annotations

from collections import defaultdict
from pathlib import Path
import json
import os
import tempfile

from openpyxl import load_workbook
from sqlalchemy import select, insert

from .db import session_scope
from .models import (
    FPTK,
    Candidate,
    Application,
    PipelineEvent,
    PositionMaster,
    MappingUser,
    SchemaConfig,
    ImportRun,
    Blacklist,
)
from .logic import (
    normalize_text,
    normalize_status,
    normalize_availability,
    excel_date,
    to_int,
    to_float,
    derive_level_number,
    derive_filter_category,
    compute_deadline,
    compute_sla_result,
    derive_recruitment_model,
)

STAGES = [
    ("Sourcing FL", "Sourcing Freelance", "Tanggal Sourcing Freelance", ""),
    ("Sourcing HR", "Sourcing HR", "Tanggal Sourcing", "Detail Keterangan Sourcing HR"),
    ("Shortlist CV", "Shortlist CV", "Tanggal Shortlist CV", "Detail Keterangan Shortlist CV"),
    ("Psikotes", "Psikotes", "Tanggal Psikotes / Cek psikotes", "Detail Keterangan Psikotes"),
    ("HR Interview", "HR Interview", "Tanggal HR Interview", "Detail Keterangan HR Interview"),
    ("Technical Test / Case Study", "Technical Test/ Case Study", "Tanggal Technical Test/ Case Study", "Detail Keterangan Technical Test/ Case Study"),
    ("Market Visit", "Market Visit", "Tanggal Market Visit", "Detail Market Visit"),
    ("User Interview", "User Interview", "Tanggal User Interview", "Detail Keterangan User Interview"),
    ("Panel Interview", "Panel Interview", "Tanggal Panel Interview", "Detail Keterangan Panel Interview"),
    ("Reference Check", "Reference Check", "Tanggal Reference Check", "Detail Keterangan Reference Check"),
    ("MCU", "MCU", "Tanggal MCU", "Detail Keterangan MCU"),
    ("Offering", "Offering", "Tanggal Offering", "Detail Keterangan Offering"),
    ("Day 1", "Day 1", "Tanggal Day 1", "Detail Keterangan Day 1"),
]
STAGE_RANK = {stage: i for i, (stage, *_rest) in enumerate(STAGES)}


def _save_temp(uploaded):
    suffix = Path(uploaded.name).suffix
    f = tempfile.NamedTemporaryFile(delete=False, suffix=suffix)
    f.write(uploaded.getbuffer())
    f.close()
    return f.name


def _header_row(ws, required, max_scan=15):
    for r in range(1, max_scan + 1):
        vals = [normalize_text(c.value) for c in ws[r]]
        if all(any((v or "").lower() == x.lower() for v in vals) for x in required):
            return r
    return 1


def _iter_records(ws, header_row):
    # Header names are constant for the whole sheet. Pre-normalize them once,
    # then each row only maps values to lowercase header keys. This matters a
    # lot for DB Sourcing (173 columns x 85k+ rows).
    headers = [normalize_text(c.value) or f"Column{c.column}" for c in ws[header_row]]
    lower_headers = [h.lower() for h in headers]
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if not any(v not in (None, "") for v in row):
            continue
        lookup = {lower_headers[i]: row[i] if i < len(row) else None for i in range(len(lower_headers))}
        yield {"__normalized_lookup__": lookup}


def _v(rec, *names):
    # Building a normalized 100+ column header map for every field lookup is
    # extremely expensive on DB Sourcing (85k+ rows). Cache it once per row.
    norm = rec.get("__normalized_lookup__")
    if norm is None:
        norm = {normalize_text(k).lower(): v for k, v in rec.items() if k and k != "__normalized_lookup__"}
        rec["__normalized_lookup__"] = norm
    for n in names:
        if n.lower() in norm:
            return norm[n.lower()]
    return None


def _norm_key(value):
    return (normalize_text(value) or "").strip()


def _fptk_key(kode, position):
    return (_norm_key(kode), _norm_key(position))


def _candidate_name_phone_key(name, phone):
    return ((_norm_key(name)).casefold(), (_norm_key(phone)).casefold())


def _candidate_email_key(email):
    return (_norm_key(email)).casefold()


def _apply_fptk_values(obj, rec, source):
    pos = normalize_text(_v(rec, "Posisi", "Position")) or obj.position
    obj.kode_pic = normalize_text(_v(rec, "Kode PIC"))
    obj.kode_angka = normalize_text(_v(rec, "Kode Angka", "KODE"))
    obj.fptk_date = excel_date(_v(rec, "FPTK Date (Real)", "Tanggal FPTK"))
    obj.fptk_date_code = excel_date(_v(rec, "FPTK Date (Kode)", "FPTK Date"))
    obj.position = pos
    obj.business_unit = normalize_text(_v(rec, "Business Unit", "PT Business Unit"))
    obj.directorate = normalize_text(_v(rec, "Direktorat", "Directorate"))
    obj.division = normalize_text(_v(rec, "Divisi", "Division"))
    obj.department = normalize_text(_v(rec, "Department"))
    obj.level_fptk = normalize_text(_v(rec, "Level FPTK", "Level FPTK (Sesuai SO)", "Level"))
    obj.level_number = to_int(_v(rec, "Level Number"), derive_level_number(obj.level_fptk))
    obj.request_reason = normalize_text(_v(rec, "Alasan Permintaan FPTK"))
    obj.category_fptk = normalize_text(_v(rec, "Category FPTK", "Category"))
    obj.pic_recruiter = normalize_text(_v(rec, "PIC Recruiter", "PIC TA", "Nama Rekruter", "Rekruter"))
    obj.status = normalize_status(_v(rec, "Status", "Status Rekrutmen"))
    obj.fptk_availability = normalize_availability(_v(rec, "FPTK Availability", "Status FPTK"), default=True)
    obj.filter_category = derive_filter_category(pos, obj.level_fptk, _v(rec, "Filter Kategorisasi FPTK"))
    obj.vacancy = to_int(_v(rec, "Vacancy", "Jumlah Permintaan"), 1) or 1
    obj.cancel_date = excel_date(_v(rec, "FPTK Cancel Date"))
    obj.offering_date = excel_date(_v(rec, "Offering Date", "Tanggal Offering"))
    obj.candidate_name = normalize_text(_v(rec, "Nama Kandidat"))
    obj.estimated_join = excel_date(_v(rec, "Estimasi Join", "Tanggal Join"))
    obj.manager_user = normalize_text(_v(rec, "User (Manager)", "User Manager"))
    obj.indirect_user = normalize_text(_v(rec, "Indirect User"))
    obj.work_location = normalize_text(_v(rec, "Lokasi Kerja", "Location"))
    obj.hr_location = normalize_text(_v(rec, "Lokasi HR"))
    obj.region = normalize_text(_v(rec, "Region"))
    obj.new_replacement = normalize_text(_v(rec, "NEW/REPLACEMENT", "Kategori"))
    obj.detail_category = normalize_text(_v(rec, "Detail Kategori"))
    obj.sla_days = to_int(_v(rec, "Jumlah SLA")) or (45 if derive_level_number(obj.level_fptk) == 4 else 30)
    obj.deadline_sla = excel_date(_v(rec, "Deadline pemenuhan SLA", "SLA Target Pemenuhan")) or compute_deadline(obj.fptk_date, obj.level_fptk)
    obj.sla_result = normalize_text(_v(rec, "Detail SLA")) or compute_sla_result(obj.status, obj.deadline_sla, obj.offering_date, obj.cancel_date)
    obj.source_file = source
    obj.remarks = normalize_text(_v(rec, "Remarks", "Recruitment Update"))


def _upsert_fptk(s, rec, source, cache=None, code_index=None):
    kode = normalize_text(_v(rec, "Kode Unik"))
    pos = normalize_text(_v(rec, "Posisi", "Position"))
    if not kode or not pos:
        return "skipped", None

    key = _fptk_key(kode, pos)
    obj = cache.get(key) if cache is not None else None
    if obj is None and cache is None:
        obj = s.scalar(select(FPTK).where(FPTK.kode_unik == kode, FPTK.position == pos))

    created = obj is None
    if created:
        obj = FPTK(kode_unik=kode, position=pos)
        s.add(obj)
        if cache is not None:
            cache[key] = obj
        if code_index is not None:
            code_index[kode].append(obj)

    _apply_fptk_values(obj, rec, source)
    return ("inserted" if created else "updated"), obj


def _upsert_reference_tables(s, wb, stats):
    # DB Kode Posisi: avoid creating the same reference rows on every re-import.
    if "DB Kode Posisi" in wb.sheetnames:
        existing = {
            (
                _norm_key(x.code), _norm_key(x.position), _norm_key(x.location),
                _norm_key(x.business_unit), _norm_key(x.division), _norm_key(x.department),
                _norm_key(x.manager_user), _norm_key(x.indirect_user), _norm_key(x.directorate), x.year,
            )
            for x in s.scalars(select(PositionMaster)).all()
        }
        ws = wb["DB Kode Posisi"]
        hr = _header_row(ws, ["POSITION"])
        for rec in _iter_records(ws, hr):
            p = normalize_text(_v(rec, "POSITION", "Position"))
            if not p:
                continue
            vals = dict(
                code=normalize_text(_v(rec, "KODE")),
                position=p,
                location=normalize_text(_v(rec, "LOCATION")),
                business_unit=normalize_text(_v(rec, "BUSINESS UNIT")),
                division=normalize_text(_v(rec, "DIVISION CHRIS")),
                department=normalize_text(_v(rec, "DEPARTMENT CHRIS")),
                manager_user=normalize_text(_v(rec, "USER (MANAGER)")),
                indirect_user=normalize_text(_v(rec, "INDIRECT USER")),
                directorate=normalize_text(_v(rec, "DIRECTORATE")),
                year=to_int(_v(rec, "YEAR")),
            )
            key = (
                _norm_key(vals["code"]), _norm_key(vals["position"]), _norm_key(vals["location"]),
                _norm_key(vals["business_unit"]), _norm_key(vals["division"]), _norm_key(vals["department"]),
                _norm_key(vals["manager_user"]), _norm_key(vals["indirect_user"]), _norm_key(vals["directorate"]), vals["year"],
            )
            if key not in existing:
                s.add(PositionMaster(**vals))
                existing.add(key)
                stats["position_master_inserted"] += 1

    if "Mapping User" in wb.sheetnames:
        existing = {
            (_norm_key(x.position), _norm_key(x.manager_user), _norm_key(x.indirect_user), _norm_key(x.directorate))
            for x in s.scalars(select(MappingUser)).all()
        }
        ws = wb["Mapping User"]
        hr = _header_row(ws, ["POSITION"])
        for rec in _iter_records(ws, hr):
            p = normalize_text(_v(rec, "POSITION"))
            if not p:
                continue
            vals = dict(
                position=p,
                manager_user=normalize_text(_v(rec, "USER (MANAGER)")),
                indirect_user=normalize_text(_v(rec, "INDIRECT USER")),
                directorate=normalize_text(_v(rec, "DIRECTORATE")),
            )
            key = (_norm_key(vals["position"]), _norm_key(vals["manager_user"]), _norm_key(vals["indirect_user"]), _norm_key(vals["directorate"]))
            if key not in existing:
                s.add(MappingUser(**vals))
                existing.add(key)
                stats["mapping_user_inserted"] += 1

    if "Schema_Config" in wb.sheetnames:
        existing = {(x.database_name.casefold(), x.field_id.casefold()): x for x in s.scalars(select(SchemaConfig)).all()}
        ws = wb["Schema_Config"]
        hr = _header_row(ws, ["Database", "Field ID", "Header Name"])
        for rec in _iter_records(ws, hr):
            dbn = normalize_text(_v(rec, "Database"))
            fid = normalize_text(_v(rec, "Field ID"))
            hn = normalize_text(_v(rec, "Header Name"))
            if not (dbn and fid and hn):
                continue
            key = (dbn.casefold(), fid.casefold())
            obj = existing.get(key)
            created = obj is None
            if created:
                obj = SchemaConfig(database_name=dbn, field_id=fid, header_name=hn)
                s.add(obj)
                existing[key] = obj
            obj.header_name = hn
            obj.position = to_int(_v(rec, "Position"))
            obj.active = normalize_availability(_v(rec, "Active"))
            obj.data_type = normalize_text(_v(rec, "Data Type"))
            obj.show_in_form = normalize_availability(_v(rec, "Show In Form"))
            obj.required = normalize_availability(_v(rec, "Required"), False)
            obj.read_only = normalize_availability(_v(rec, "Read Only"), False)
            obj.searchable = normalize_availability(_v(rec, "Searchable"), False)
            obj.control_type = normalize_text(_v(rec, "Control Type"))
            obj.dropdown_values = normalize_text(_v(rec, "Dropdown Values"))
            obj.rule = normalize_text(_v(rec, "Rule"))
            obj.aliases = normalize_text(_v(rec, "Aliases"))
            obj.key_field = normalize_availability(_v(rec, "Key Field"), False)
            obj.custom_field = normalize_availability(_v(rec, "Custom Field"), False)
            obj.canonical_header = normalize_text(_v(rec, "Canonical Header"))
            stats["schema_inserted" if created else "schema_updated"] += 1

    if "Blacklist Candidate" in wb.sheetnames:
        existing = {
            ((_norm_key(x.candidate_name)).casefold(), (_norm_key(x.reason)).casefold())
            for x in s.scalars(select(Blacklist)).all()
        }
        ws = wb["Blacklist Candidate"]
        hr = _header_row(ws, ["Nama Kandidat"])
        for rec in _iter_records(ws, hr):
            name = normalize_text(_v(rec, "Nama Kandidat", "Nama"))
            if not name:
                continue
            reason = normalize_text(_v(rec, "Alasan Tidak Proceed", "Reason"))
            key = (name.casefold(), (_norm_key(reason)).casefold())
            if key in existing:
                continue
            s.add(Blacklist(candidate_name=name, reason=reason, active=True))
            existing.add(key)
            stats["blacklist_inserted"] += 1


def _update_candidate_from_record(c, rec):
    # Fill/refresh structured fields while preserving richer existing data when the source cell is blank.
    mappings = [
        ("domicile", "Domisili", normalize_text),
        ("education_level", "Jenjang Pendidikan", normalize_text),
        ("university_top10", "Nama Universitas/Sekolah (TOP 10)", normalize_text),
        ("university_other", "Nama Universitas/Sekolah Lainnya", normalize_text),
        ("major", "Jurusan", normalize_text),
        ("graduation_year", "Tahun Lulus", to_int),
        ("gpa", "IPK", to_float),
        ("english_score", "Skor Bahasa Inggris", to_float),
        ("last_position", "Last Position", normalize_text),
        ("last_tenure", "Last Tenure", normalize_text),
        ("last_company", "Last Company", normalize_text),
        ("total_tenure", "Total Tenure", normalize_text),
    ]
    for attr, col, conv in mappings:
        raw = _v(rec, col)
        if raw not in (None, ""):
            val = conv(raw)
            if val not in (None, ""):
                setattr(c, attr, val)
    raw_fmcg = _v(rec, "Pernah di FMCG?")
    if raw_fmcg not in (None, ""):
        c.fmcg_experience = normalize_availability(raw_fmcg, False)


def import_workbook(uploaded, source_type="recruiter", progress_callback=None):
    """Import a legacy recruiter/master workbook into the central database.

    The import is deliberately idempotent for core entities:
    - FPTK is UPSERTed by (Kode Unik, Position)
    - Candidate is reused by email, then by (name, phone)
    - Application is reused by (FPTK, Candidate)
    - Pipeline event is reused by (Application, Stage, Event Date)

    Important: SessionLocal uses autoflush=False. We therefore explicitly flush the
    FPTK phase before DB Sourcing is read; otherwise DB Sourcing cannot see newly
    queued FPTKs and may attempt to insert the same unique key a second time.
    """
    path = _save_temp(uploaded)
    try:
        wb = load_workbook(path, read_only=True, data_only=True, keep_vba=False)
        counts = {"inserted": 0, "updated": 0, "skipped": 0}
        stats = defaultdict(int)

        def report(stage, current, total, message):
            if progress_callback:
                progress_callback(stage, current, total, message)

        with session_scope() as s:
            # Preload FPTK cache so duplicates are handled both against the DB and
            # within the same workbook before any SQL flush occurs.
            existing_fptk = s.scalars(select(FPTK)).all()
            fptk_cache = {_fptk_key(x.kode_unik, x.position): x for x in existing_fptk}
            code_index = defaultdict(list)
            for x in existing_fptk:
                code_index[x.kode_unik].append(x)

            if "FPTK" in wb.sheetnames:
                ws = wb["FPTK"]
                hr = _header_row(ws, ["Kode Unik", "Posisi"])
                for rec in _iter_records(ws, hr):
                    outcome, _ = _upsert_fptk(s, rec, uploaded.name, fptk_cache, code_index)
                    counts[outcome] += 1

                # CRITICAL FIX: make imported FPTKs visible to subsequent SELECTs.
                s.flush()
                report("FPTK", counts["inserted"] + counts["updated"], max(ws.max_row - hr, 1), "FPTK imported")

            _upsert_reference_tables(s, wb, stats)
            s.flush()

            if "DB Sourcing" in wb.sheetnames:
                # Batch importer for the large DB Sourcing sheet. It keeps only a
                # small chunk of candidate/application/event payloads in memory,
                # while global identity maps make the import idempotent across
                # chunks and across repeated imports.
                candidate_alias = {}
                for cid, cname, cemail, cphone in s.execute(
                    select(Candidate.id, Candidate.name, Candidate.email, Candidate.phone)
                ):
                    if cemail:
                        candidate_alias[("e", _candidate_email_key(cemail))] = cid
                    candidate_alias[("np",) + _candidate_name_phone_key(cname, cphone)] = cid

                app_id_by_db_key = {}
                app_current_stage = {}
                for aid, fid, cid, current_stage in s.execute(
                    select(Application.id, Application.fptk_id, Application.candidate_id, Application.current_stage)
                ):
                    db_key = (fid, cid)
                    app_id_by_db_key[db_key] = aid
                    app_current_stage[db_key] = current_stage

                existing_event_keys = set(
                    s.execute(
                        select(PipelineEvent.application_id, PipelineEvent.stage, PipelineEvent.event_date)
                    ).all()
                )

                temp_seq = 0
                chunk_new_candidates = {}
                chunk_candidate_updates = {}
                chunk_apps = {}
                chunk_events = {}
                chunk_temp_aliases = defaultdict(set)
                CHUNK_SIZE = 20000

                def candidate_profile(rec, name, email, phone, *, for_insert=False):
                    raw_fmcg = _v(rec, "Pernah di FMCG?")
                    fmcg = None if raw_fmcg in (None, "") else normalize_availability(raw_fmcg, False)
                    out = {
                        "name": name,
                        "email": email,
                        "phone": phone,
                        "domicile": normalize_text(_v(rec, "Domisili")),
                        "education_level": normalize_text(_v(rec, "Jenjang Pendidikan")),
                        "university_top10": normalize_text(_v(rec, "Nama Universitas/Sekolah (TOP 10)")),
                        "university_other": normalize_text(_v(rec, "Nama Universitas/Sekolah Lainnya")),
                        "major": normalize_text(_v(rec, "Jurusan")),
                        "graduation_year": to_int(_v(rec, "Tahun Lulus")),
                        "gpa": to_float(_v(rec, "IPK")),
                        "english_score": to_float(_v(rec, "Skor Bahasa Inggris")),
                        "last_position": normalize_text(_v(rec, "Last Position")),
                        "last_tenure": normalize_text(_v(rec, "Last Tenure")),
                        "last_company": normalize_text(_v(rec, "Last Company")),
                        "total_tenure": normalize_text(_v(rec, "Total Tenure")),
                        "fmcg_experience": False if (for_insert and fmcg is None) else fmcg,
                    }
                    return out

                def merge_profile(target, incoming):
                    for k, v in incoming.items():
                        if v not in (None, ""):
                            target[k] = v

                def resolve_candidate_id(token, token_to_id):
                    if isinstance(token, int):
                        return token
                    return token_to_id.get(token)

                def flush_sourcing_chunk():
                    nonlocal chunk_new_candidates, chunk_candidate_updates, chunk_apps, chunk_events, chunk_temp_aliases
                    if not (chunk_new_candidates or chunk_candidate_updates or chunk_apps or chunk_events):
                        return

                    # 1) Candidate updates + inserts.
                    token_to_id = {}
                    if chunk_new_candidates:
                        tokens = list(chunk_new_candidates.keys())
                        rows = []
                        for token in tokens:
                            row = dict(chunk_new_candidates[token])
                            if row.get("fmcg_experience") is None:
                                row["fmcg_experience"] = False
                            rows.append(row)
                        result = s.execute(
                            insert(Candidate).returning(
                                Candidate.id, Candidate.name, Candidate.email, Candidate.phone
                            ),
                            rows,
                        )
                        returned = list(result)
                        # RETURNING order follows insert order for SQLAlchemy's
                        # insertmanyvalues on supported SQLite/PostgreSQL. We still
                        # register both aliases from returned values afterward.
                        for token, (cid, cname, cemail, cphone) in zip(tokens, returned):
                            token_to_id[token] = cid
                            # Replace every alias learned for this temporary token,
                            # not only the final row's email/name-phone pair.
                            for alias in chunk_temp_aliases.get(token, set()):
                                candidate_alias[alias] = cid
                            if cemail:
                                candidate_alias[("e", _candidate_email_key(cemail))] = cid
                            candidate_alias[("np",) + _candidate_name_phone_key(cname, cphone)] = cid
                        stats["candidate_db_inserted"] += len(returned)

                    # 2) Applications.
                    app_rows = []
                    app_updates = []
                    for (fid, token), app in chunk_apps.items():
                        cid = resolve_candidate_id(token, token_to_id)
                        if cid is None:
                            stats["application_no_candidate"] += 1
                            continue
                        db_key = (fid, cid)
                        aid = app_id_by_db_key.get(db_key)
                        new_stage = app.get("current_stage")
                        if aid is None:
                            app_rows.append({
                                "fptk_id": fid,
                                "candidate_id": cid,
                                "recruiter": app.get("recruiter"),
                                "model": app.get("model"),
                                "status": app.get("status") or "ACTIVE",
                                "current_stage": new_stage,
                            })
                        else:
                            stats["application_reused"] += 1
                            old_stage = app_current_stage.get(db_key)
                            if STAGE_RANK.get(new_stage, -1) > STAGE_RANK.get(old_stage, -1):
                                app_updates.append({"id": aid, "current_stage": new_stage})
                                app_current_stage[db_key] = new_stage

                    if app_rows:
                        result = s.execute(
                            insert(Application).returning(
                                Application.id, Application.fptk_id, Application.candidate_id, Application.current_stage
                            ),
                            app_rows,
                        )
                        returned_apps = list(result)
                        for aid, fid, cid, current_stage in returned_apps:
                            db_key = (fid, cid)
                            app_id_by_db_key[db_key] = aid
                            app_current_stage[db_key] = current_stage
                        stats["application_inserted"] += len(returned_apps)
                    if app_updates:
                        s.bulk_update_mappings(Application, app_updates)
                        stats["application_updated"] += len(app_updates)

                    # 3) Pipeline events.
                    event_rows = []
                    for (fid, token, stage, dt), ev in chunk_events.items():
                        cid = resolve_candidate_id(token, token_to_id)
                        if cid is None:
                            continue
                        app_id = app_id_by_db_key.get((fid, cid))
                        if app_id is None:
                            stats["event_no_application"] += 1
                            continue
                        db_ev_key = (app_id, stage, dt)
                        if db_ev_key in existing_event_keys:
                            stats["event_reused"] += 1
                            continue
                        existing_event_keys.add(db_ev_key)
                        event_rows.append({
                            "application_id": app_id,
                            "stage": stage,
                            "result": ev.get("result"),
                            "detail": ev.get("detail"),
                            "event_date": dt,
                            "created_by": "legacy-import",
                        })
                    if event_rows:
                        s.execute(insert(PipelineEvent), event_rows)
                        stats["event_inserted"] += len(event_rows)

                    s.flush()
                    chunk_new_candidates = {}
                    chunk_candidate_updates = {}
                    chunk_apps = {}
                    chunk_events = {}
                    chunk_temp_aliases = defaultdict(set)

                ws = wb["DB Sourcing"]
                hr = _header_row(ws, ["Kode Unik", "Nama"])
                source_rows = 0

                for rec in _iter_records(ws, hr):
                    source_rows += 1
                    kode = normalize_text(_v(rec, "Kode Unik"))
                    name = normalize_text(_v(rec, "Nama"))
                    pos = normalize_text(_v(rec, "Posisi"))
                    if not kode or not name:
                        stats["sourcing_skipped"] += 1
                        if source_rows % CHUNK_SIZE == 0:
                            flush_sourcing_chunk()
                        continue

                    f = fptk_cache.get(_fptk_key(kode, pos)) if pos else None
                    if not f:
                        matches = code_index.get(kode, [])
                        if len(matches) == 1:
                            f = matches[0]
                    if not f:
                        stats["sourcing_no_fptk"] += 1
                        if source_rows % CHUNK_SIZE == 0:
                            flush_sourcing_chunk()
                        continue

                    email = normalize_text(_v(rec, "Email"))
                    phone = normalize_text(_v(rec, "Nomor HP"))
                    email_alias = ("e", _candidate_email_key(email)) if email else None
                    np_alias = ("np",) + _candidate_name_phone_key(name, phone)
                    token_email = candidate_alias.get(email_alias) if email_alias else None
                    token_np = candidate_alias.get(np_alias)

                    if token_email is not None and token_np is not None and token_email != token_np:
                        token = token_email
                        stats["candidate_identity_conflict"] += 1
                    else:
                        token = token_email if token_email is not None else token_np

                    if token is None:
                        temp_seq += 1
                        token = ("new", temp_seq)
                        candidate_alias[np_alias] = token
                        chunk_temp_aliases[token].add(np_alias)
                        if email_alias:
                            candidate_alias[email_alias] = token
                            chunk_temp_aliases[token].add(email_alias)
                        chunk_new_candidates[token] = candidate_profile(rec, name, email, phone, for_insert=True)
                        stats["candidate_inserted"] += 1
                    elif isinstance(token, int):
                        # Candidate already exists in the central DB. Legacy import
                        # is a migration/compatibility layer, so do not rewrite the
                        # same candidate profile on every sourcing row; that turns
                        # large imports into tens of thousands of UPDATE statements.
                        candidate_alias[np_alias] = token
                        if email_alias:
                            candidate_alias[email_alias] = token
                        stats["candidate_reused"] += 1
                    else:
                        merge_profile(
                            chunk_new_candidates.setdefault(token, candidate_profile(rec, name, email, phone, for_insert=True)),
                            candidate_profile(rec, name, email, phone, for_insert=True),
                        )
                        candidate_alias[np_alias] = token
                        chunk_temp_aliases[token].add(np_alias)
                        if email_alias:
                            candidate_alias[email_alias] = token
                            chunk_temp_aliases[token].add(email_alias)
                        stats["candidate_reused"] += 1

                    app_key = (f.id, token)
                    app = chunk_apps.get(app_key)
                    if app is None:
                        app = {
                            "recruiter": normalize_text(_v(rec, "Rekruter")) or f.pic_recruiter,
                            "model": normalize_text(_v(rec, "Model Rekrutmen")) or derive_recruitment_model(
                                f.position, f.level_fptk, f.filter_category
                            ),
                            "status": "ACTIVE",
                            "current_stage": None,
                        }
                        chunk_apps[app_key] = app

                    latest = None
                    for stage, status_col, date_col, detail_col in STAGES:
                        result = normalize_text(_v(rec, status_col))
                        dt = excel_date(_v(rec, date_col))
                        detail = normalize_text(_v(rec, detail_col)) if detail_col else None
                        if result or dt or detail:
                            ev_key = (f.id, token, stage, dt)
                            existing_ev = chunk_events.get(ev_key)
                            if existing_ev is None:
                                chunk_events[ev_key] = {"result": result, "detail": detail}
                            else:
                                if result:
                                    existing_ev["result"] = result
                                if detail:
                                    existing_ev["detail"] = detail
                            latest = stage
                    if latest and STAGE_RANK.get(latest, -1) >= STAGE_RANK.get(app.get("current_stage"), -1):
                        app["current_stage"] = latest

                    if source_rows % CHUNK_SIZE == 0:
                        flush_sourcing_chunk()
                        report("DB Sourcing", source_rows, max(ws.max_row - hr, 1), f"DB Sourcing {source_rows:,} rows")

                flush_sourcing_chunk()
                report("DB Sourcing", source_rows, max(ws.max_row - hr, 1), f"DB Sourcing {source_rows:,} rows complete")
                stats["sourcing_rows_read"] = source_rows

            notes = {
                "message": "Legacy workbook import (idempotent UPSERT)",
                **dict(stats),
            }
            s.add(
                ImportRun(
                    source_file=uploaded.name,
                    source_type=source_type,
                    inserted=counts["inserted"],
                    updated=counts["updated"],
                    skipped=counts["skipped"],
                    notes=json.dumps(notes, ensure_ascii=False),
                )
            )

        return {**counts, **dict(stats)}
    finally:
        try:
            os.unlink(path)
        except OSError:
            pass
